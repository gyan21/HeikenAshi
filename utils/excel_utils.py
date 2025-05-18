import os
from openpyxl import Workbook, load_workbook

EXCEL_LOG_FILE = "trade_log.xlsx"

def save_trade_to_excel(trade_info, excel_file=EXCEL_LOG_FILE):
    headers = [
        "date", "spread", "type", "symbol", "sell_strike", "buy_strike", "expiry",
        "open_price", "close_price", "profit", "profit_pct", "status", "close_reason", "quantity"
    ]
    if not os.path.exists(excel_file):
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        wb.save(excel_file)

    wb = load_workbook(excel_file)
    ws = wb.active

    # Ensure all headers are present in trade_info
    row = [trade_info.get(h, "") for h in headers]
    ws.append(row)
    wb.save(excel_file)